import datetime
from math import *
from openpyxl import *
import holidays


#These dictionaries help get all the dates correct
day_in_month_dict = {'January': 31, 'February': 28, 'March': 31, 'April': 30, 'May': 31, 'June': 30, 'July': 31,
                     'August': 31, 'September': 30, 'October': 31, 'November': 30, 'December': 31}
day_in_month_dict_num = {1: 31, 2: 28, 3: 31, 4: 30, 5: 31, 6: 30, 7: 31,
                         8: 31, 9: 30, 10: 31, 11: 30, 12: 31}
month_dict = {'January': 1, 'February': 2, 'March': 3, 'April': 4, 'May': 5, 'June': 6, 'July': 7, 'August': 8, 'September': 9, 'October': 10, 'November': 11, 'December': 12}
days_of_week = ['Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
num_vs_mon_dict = {1: 'Jan', 2: 'Feb', 3: 'Mar', 4: 'Apr', 5: 'May', 6: 'Jun', 7: 'Jul', 8: 'Aug', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dec'}
#lists for the amount of fasting per day
the_16_to_24_list = ['8-24hr', '8-12hr', '8-12hr', '8-16hr', '8-17hr', '8-18hr', '8-24hr',
                     '8-16hr', '8-12hr', '8-16hr', '17hr', '18-20hr', '19-21hr', '8-24hr',
                     '8-24hr', '8-16hr', '16-20hr', '16-24hr', '18-24hr', '20-24hr', '8-24hr',
                     '8-24hr', '8-16hr', '16-20hr',	'20-24 hr',	'21-24hr', '22-24hr', '8-24hr',
                     '8-24hr', '8-16hr', '16-24hr', '18-24hr', '20-24hr', '22-24hr','8-24hr',
                     '8-24hr', '16-24hr', '20-26hr', '20-36hr', '***', '16-25hr', '8-24hr',
                     '8-24hr', '16-24hr', '16-24hr', '18-24hr',	'24-36hr', '***', '8-24hr']
the_24_to_36_list = ['8-16hr', '8-16 hr', '8-12hr', '8-16hr', '8-17hr',	'8-18hr', '8-24hr',
                     '8-16hr', '8-16hr', '8-16hr', '17hr', '18-20hr', '19-21hr', '8-24hr',
                     '8-24hr', '8-16hr', '16-20hr',	'16-24hr', '18-24hr', '20-24hr', '8-24hr',
                     '8-24 hr',	'8-16hr', '20-24hr', '20-36hr',	'***', '24-36hr', '***',
                     '8-24hr', '16-24hr', '22-36hr', '***',	'24-36hr', '***', '8-24hr',
                     '8-24hr', '16-24hr', '24-36hr', '***',	'24-42hr', '***', '8-24hr',
                     '8-24hr', '8-24hr', '16-24hr', '18-24hr', '24-72hr', '***', '***']

the_24_to_48_list = ['8-16hr', '8-16hr', '8-12hr', '8-16hr', '8-17hr', '8-18hr', '8-24hr',
                     '8-16hr', '8-16hr', '8-16hr', '17hr', '18-20hr', '19-21hr', '8-24hr',
                     '8-24hr',	'8-16hr', '16-20hr', '16-24hr',	'18-24hr', '20-24hr', '8-24hr',
                     '8-24 hr',	'8-16hr', '20-24hr', '20-36hr',	'***', '24-36hr', '***',
                     '8-24hr', '16-24hr', '24-42hr', '***', '24-48hr', '***', '8-24hr',
                     '8-24hr', '16-24hr', '24-72hr', '***',	'***', '16-24hr', '8-24hr',
                     '8-24hr', '8-24hr', '16-24hr',	'18-24hr', '24-72hr', '***', '***']

the_36_to_48_list = ['8-24 hr', '8-24hr', '8-18hr', '16-24hr', '18-36hr', '***', '8-24 hr',
                     '8-24hr', '8-24hr', '16-24hr', '20-24hr', '24-36hr',	'***', '8-24hr',
                     '8-24hr', '8-24hr', '22-24hr',	'24-36hr', '***', '36-48hr', '***',
                     '8-24hr',	'8-24 hr', '8-24hr', '20-26hr',	'24-42hr', '***', '8-24hr',
                     '8-24hr', '8-24hr', '36-48hr', '***', '36-48hr', '***', '8-24hr',
                     '8-24hr',	'8-24hr', '24hr', '36-96hr', '***', '***', '***',
                     '8-24hr', '16-24hr', '36-72hr', '***',	'***', '8-24hr', '8-24hr']
the_3_day_to_4_day_list = ['8-24hr', '8-24hr', '16-24hr', '16-24hr', '18-24 hr', '20-24hr',	'8-24hr',
                           '8-24hr', '16-24hr', '16-24hr', '24-36hr',	'***', '24-36hr', '***',
                           '8-24hr', '16-24hr',	'20-24hr', '3 day',	'***', '***', '8-24hr',
                           '8-24hr', '16-24hr',	'20-24hr', '20-24hr', '3 day', '***', '***',
                           '8-24hr', '16-24hr',	'20-24hr', '3 day', '***', '***', '8-24hr',
                           '8-24hr', '16-24hr',	'20-26hr', '4 day', '***', '***', '***',
                           '8-24hr', '16-24hr',	'20-24hr', '20-24hr', '3 day', '***', '***']

the_72_to_7_day_list = ['8-24hr', '8-24hr', '12-24hr', '36hr', '***', '20-24hr', '8-24hr',
                        '8-16hr', '16-24hr', '18-24hr', '36hr', '***', '42hr', '***',
                        '16-24hr', '3 to 7 day', '***',	'***', '***', '***', '***',
                        '***', '8-24hr', '16-24hr', '3-4 day', '***', '***', '***',
                        '8-24hr', '16-24hr', '36-42hr',	'***', '36-48hr', '***', '8-24hr',
                        '8-24hr', '3 to 7 day',	'***',	'***', '***', '***', '***',
                        '***', '8-24hr', '16-24hr', '24-36hr', '***', '8-24hr',	'8-24hr']
#This is for making the cells different colors
grayFill = styles.fills.PatternFill(start_color='00808080',
                   end_color='00808080',
                   fill_type='solid')
darkgrayFill = styles.fills.PatternFill(start_color='00333333',
                   end_color='00333333',
                   fill_type='solid')
darkpinkFill = styles.fills.PatternFill(start_color='00993366',
                   end_color='00993366',
                   fill_type='solid')
yellowFill = styles.fills.PatternFill(start_color='00FFFF00',
                   end_color='00FFFF00',
                   fill_type='solid')
dirtyyellowFill = styles.fills.PatternFill(start_color='00FFCC00',
                   end_color='00FFCC00',
                   fill_type='solid')
purpleFill = styles.fills.PatternFill(start_color='00800080',
                   end_color='00800080',
                   fill_type='solid')
lightpurpleFill = styles.fills.PatternFill(start_color='00CC99FF',
                   end_color='00CC99FF',
                   fill_type='solid')
neongreenFill = styles.fills.PatternFill(start_color='0000FF00',
                   end_color='0000FF00',
                   fill_type='solid')
lightgreenFill = styles.fills.PatternFill(start_color='0099CC00',
                   end_color='0099CC00',
                   fill_type='solid')
darkgreenFill = styles.fills.PatternFill(start_color='00008000',
                   end_color='00008000',
                   fill_type='solid')
lightblueFill = styles.fills.PatternFill(start_color='0000CCFF',
                   end_color='0000CCFF',
                   fill_type='solid')
redbrownFill = styles.fills.PatternFill(start_color='00993300',
                   end_color='00993300',
                   fill_type='solid')


def isLeap(year):
    """
    Checks if the year is a leap year or not
    """
    if (year % 100) == 0:
        if (year % 400) == 0:
            return True
        else:
            return False
    if (year % 4) == 0:
        return True
    else:
        return False



def main():
    First_name = input("Enter the First Name: ")
    Last_name = input("\nEnter the Last Name: ")
    start_Month = input("\nEnter a Month: ")
    start_date = int(input("\nEnter the date: "))
    height = int(input("\nHow tall are you in inches? "))
    which_fast = int(input("\n1. 16 to 24 hours\n2. 24 to 36 hours\n3. 24 to 48 hours\n4. 36 to 48 hours\n5. 3 to 4 days\n6. 3 to 7 days\nEnter the Number the corrapotes to the desired fast: "))
    vaca_date_start = 0
    vaca_month_end = 'Skadoosh'
    vaca_date_end = 0
    vaca_month_start = input("\nEnter the month you want to start your vacation, if you don't want to start one enter \'Skadoosh\': ")
    if vaca_month_start != 'Skadoosh':
        vaca_date_start = int(input("\nEnter vacation start date: "))
        vaca_month_end = input("\nEnter the month you want to end your vacation, if you only want one day off enter \'Skadoosh\', of if the vacation is within the same month please it again: ")
        if vaca_month_end != 'Skadoosh':
            vaca_date_end = int(input("\nEnter vacation end date: "))
    vaca_month_start_2 = input(
        "\nEnter the month you want to start your 2nd vacation, if you don't want to start one enter \'Skadoosh\': ")
    if vaca_month_start_2 != 'Skadoosh':
        vaca_date_start_2 = int(input("\nEnter the 2nd vacation start date: "))
        vaca_month_end_2 = input(
            "\nEnter the month you want to end your 2nd vacation, if you only want one day off enter \'Skadoosh\', of if the vacation is within the same month please it again: ")
        if vaca_month_end_2 != 'Skadoosh':
            vaca_date_end_2 = int(input("\nEnter the 2nd vacation end date: "))
    vaca_month_start_3 = input(
        "\nEnter the month you want to start your 3rd vacation, if you don't want to start one enter \'Skadoosh\': ")
    if vaca_month_start_3 != 'Skadoosh':
        vaca_date_start_3 = int(input("\nEnter the 3rd vacation start date: "))
        vaca_month_end_3 = input(
            "\nEnter the month you want to end your 3rd vacation, if you only want one day off enter \'Skadoosh\', of if the vacation is within the same month please it again: ")
        if vaca_month_end_3 != 'Skadoosh':
            vaca_date_end_3 = int(input("\nEnter the 3rd vacation end date: "))


    six_week(First_name, Last_name, start_Month, start_date, height, which_fast, vaca_month_start, vaca_date_start, vaca_month_end, vaca_date_end, vaca_month_start_2, vaca_date_start_2, vaca_month_end_2, vaca_date_end_2, vaca_month_start_3, vaca_date_start_3, vaca_month_end_3, vaca_date_end_3)


def six_week(First_name, Last_name, start_Month, start_date, height, which_fast, vaca_month_start='Skadoosh', vaca_date_start=0, vaca_month_end='Skadoosh', vaca_date_end=0, vaca_month_start_2='Skadoosh', vaca_date_start_2=0, vaca_month_end_2='Skadoosh', vaca_date_end_2=0, vaca_month_start_3='Skadoosh', vaca_date_start_3=0, vaca_month_end_3='Skadoosh', vaca_date_end_3=0):
    """
    Creates a Microsoft Excel document that shows how long a person should fast based on inputted factors
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"{First_name}-{Last_name} 6 week plan"
    stop_weight = floor((17 * (height ** 2)) / 703)
    year = datetime.date.today().year
    the_list = []
    vaca_list = []
    counter = 0
    if which_fast == 1:
        for i in range(49):
            the_list.append(the_16_to_24_list[i])
    elif which_fast == 2:
        for i in range(49):
            the_list.append(the_24_to_36_list[i])
    elif which_fast == 3:
        for i in range(49):
            the_list.append(the_24_to_48_list[i])
    elif which_fast == 4:
        for i in range(49):
            the_list.append(the_36_to_48_list[i])
    elif which_fast == 5:
        for i in range(49):
            the_list.append(the_3_day_to_4_day_list[i])
    elif which_fast == 6:
        for i in range(49):
            the_list.append(the_72_to_7_day_list[i])
    month_num = month_dict[start_Month]
    if datetime.datetime(year, month_num, start_date).weekday() != 6:
        while datetime.datetime(year, month_num, start_date).weekday() != 6:
            start_date -= 1
            if start_date == 0:
                month_num -= 1
                if month_num == 0:
                    month_num = 12
                    year -= 1
                    start_date = day_in_month_dict_num[month_num]
                elif month_num == 2:
                    if isLeap(year) == True:
                        start_date = 29
                    else:
                        start_date = 28
                else:
                    start_date = day_in_month_dict_num[month_num]
    year = datetime.date.today().year
    vaca_year = year
    og_start_date = start_date
    if vaca_month_start != 'Skadoosh':
        month_start_num = month_dict[vaca_month_start]
        if vaca_month_end != 'Skadoosh':
            month_end_num = month_dict[vaca_month_end]
            while f'{vaca_date_start}-{num_vs_mon_dict[month_start_num]}' != f'{vaca_date_end + 1}-{num_vs_mon_dict[month_end_num]}':
                vaca_list.append(f'{vaca_date_start}-{num_vs_mon_dict[month_start_num]}')
                vaca_date_start += 1
                if month_start_num == 2:
                    if (isLeap(vaca_year)) == True:
                        if vaca_date_start == 30:
                            month_start_num += 1
                            vaca_date_start = 1
                    else:
                        if vaca_date_start == 29:
                            month_start_num += 1
                            vaca_date_start = 1
                if vaca_date_start == 31:
                    if day_in_month_dict_num[month_start_num] == 30:
                        month_start_num += 1
                        vaca_date_start = 1
                if vaca_date_start == 32:
                    month_start_num += 1
                    if month_start_num == 13:
                        month_start_num = 1
                        vaca_year += 1
                    vaca_date_start = 1
        else:
            vaca_list.append(f'{vaca_date_start}-{num_vs_mon_dict[month_dict[vaca_month_start]]}')
    vaca_year_2 = year
    if vaca_month_start_2 != 'Skadoosh':
        month_start_num_2 = month_dict[vaca_month_start_2]
        if vaca_month_end_2 != 'Skadoosh':
            month_end_num_2 = month_dict[vaca_month_end_2]
            while f'{vaca_date_start_2}-{num_vs_mon_dict[month_start_num_2]}' != f'{vaca_date_end_2 + 1}-{num_vs_mon_dict[month_end_num_2]}':
                vaca_list.append(f'{vaca_date_start_2}-{num_vs_mon_dict[month_start_num_2]}')
                vaca_date_start_2 += 1
                if month_start_num_2 == 2:
                    if (isLeap(vaca_year)) == True:
                        if vaca_date_start_2 == 30:
                            month_start_num_2 += 1
                            vaca_date_start_2 = 1
                    else:
                        if vaca_date_start_2 == 29:
                            month_start_num_2 += 1
                            vaca_date_start_2 = 1
                if vaca_date_start == 31:
                    if day_in_month_dict_num[month_start_num_2] == 30:
                        month_start_num_2 += 1
                        vaca_date_start_2 = 1
                if vaca_date_start_2 == 32:
                    month_start_num_2 += 1
                    if month_start_num_2 == 13:
                        month_start_num_2 = 1
                        vaca_year_2 += 1
                    vaca_date_start_2 = 1
        else:
            vaca_list.append(f'{vaca_date_start_2}-{num_vs_mon_dict[month_dict[vaca_month_start_2]]}')

        vaca_year_3 = year
        if vaca_month_start_3 != 'Skadoosh':
            month_start_num_3 = month_dict[vaca_month_start_2]
            if vaca_month_end_3 != 'Skadoosh':
                month_end_num_3 = month_dict[vaca_month_end_2]
                while f'{vaca_date_start_3}-{num_vs_mon_dict[month_start_num_3]}' != f'{vaca_date_end_3 + 1}-{num_vs_mon_dict[month_end_num_3]}':
                    vaca_list.append(f'{vaca_date_start_3}-{num_vs_mon_dict[month_start_num_3]}')
                    vaca_date_start_3 += 1
                    if month_start_num_3 == 2:
                        if (isLeap(vaca_year)) == True:
                            if vaca_date_start_3 == 30:
                                month_start_num_3 += 1
                                vaca_date_start_3 = 1
                        else:
                            if vaca_date_start_3 == 29:
                                month_start_num_3 += 1
                                vaca_date_start_3 = 1
                    if vaca_date_start == 31:
                        if day_in_month_dict_num[month_start_num_3] == 30:
                            month_start_num_3 += 1
                            vaca_date_start_3 = 1
                    if vaca_date_start_3 == 32:
                        month_start_num_3 += 1
                        if month_start_num_3 == 13:
                            month_start_num_3 = 1
                            vaca_year_3 += 1
                        vaca_date_start_3 = 1
            else:
                vaca_list.append(f'{vaca_date_start_3}-{num_vs_mon_dict[month_dict[vaca_month_start_3]]}')

    #lines 140-211 Input info into cells that are constant regardless of inputted factors
    for i in range (3):
        ws.merge_cells(f'A{4 + i}:K{4 + i}')
    for i in range(3):
        ws.merge_cells(f'A{7 + i}:I{7 + i}')
    for i in range(80):
        ws.merge_cells(f'J{11 + i}:M{11+i}')
    ws.merge_cells('D43:E43')
    ws.merge_cells('A1:G3')
    ws['A1'].value = '6 Week Plan'
    ws['A1'].font = styles.Font(size=40, underline='single')
    ws['A4'].value = 'To My Family and Myself: Fasting is Science Based and has been shown to be safe. I\'m doing this 6 week plan for my Health.'
    ws['A4'].fill = neongreenFill
    ws['A5'].value = 'It is okay that I don\'t do this perfectly, but I need support for My Success.  I will listen to my body and stop on days I don\'t feel well.'
    ws['A5'].fill = neongreenFill
    ws['A6'].value = 'In addition, to keep this safe, I have a stop weight that, if I reach, I promise to stop.'
    ws['A6'].fill = neongreenFill
    ws['A7'].value = 'While I have a personal weight goal I am trying to reach, my starting weight is already at a point'
    ws['A7'].fill = lightpurpleFill
    ws['A8'].value = 'that someone else would do just about anything to be at. I can do this! And I can do this safely.'
    ws['A8'].fill = lightpurpleFill
    ws['A9'].value = 'LETS GO!!!'
    ws['A9'].font = styles.Font(bold=True, underline='single')
    ws['A9'].fill = lightpurpleFill
    ws['J12'].value = 'Goal: '
    ws['J12'].fill = lightgreenFill
    ws['J14'].value = 'Maintain weight over 6 weeks'
    ws['J15'].value = 'Lose XX pounds'
    ws['J16'].value = 'Current Fasting Record: '
    ws['J17'].value = 'Goal Fast: '
    ws['J18'].value = 'Stop Eating Time: '
    ws['J19'].value = 'Take before pic/measurements'
    ws['J21'].value = 'Avoid Sugar/Sweetners Days'
    ws['J21'].fill = purpleFill
    ws['J22'].value = 'This will make it easier to Fast longer'
    ws['J24'].value = 'Special Days: '
    ws['J24'].fill = lightblueFill
    ws['J25'].value = 'Vacation Days'
    ws['J26'].value = 'Celebrations: Fast until Party'
    ws['J28'].value = 'Food Tracking'
    ws['J28'].fill = lightgreenFill
    ws['J29'].value = 'What you ate, not amounts'
    ws['J30'].value = 'Try a Fasting Fuel: Nuts, Olives,'
    ws['J31'].value = 'Sliced Pickles'
    ws['J32'].value = 'Take a Multivitamin'
    ws['J34'].value = 'Exercise example: C-30 mins'
    ws['J34'].fill = darkpinkFill
    ws['J35'].value = 'Go Easier on Longer Fast Days'
    ws['J35'].font = styles.Font(bold=True, underline='single')
    ws['J36'].value = 'W=Weights'
    ws['J37'].value = 'C=Cardio'
    ws['J38'].value = 'BK= Bad Knees 5 K plan'
    ws['J40'].value = 'Meds:'
    ws['J40'].fill = darkgreenFill
    ws['J41'].value = 'Use a Fasting Fuel if can’t take with just'
    ws['J42'].value = 'Water'
    ws['J44'].value = f'Stop Weight: {stop_weight}'
    ws['J44'].font = styles.Font(color="00FF0000")
    ws['J44'].fill = darkgrayFill
    ws['J45'].value = 'Health risks increase with BMI Below 17'
    ws['J45'].font = styles.Font(color="00FF0000")
    ws['J45'].fill = darkgrayFill
    ws['J47'].value = 'Weight'
    ws['J48'].value = 'Weigh-in Wednesday (weigh weekly)'
    ws['J49'].value = 'Daily weights will vary, this is normal'
    ws['J50'].value = '***'
    ws['J50'].fill = yellowFill
    ws['J51'].value = 'Keep Going or Break Fast when Hungry'
    ws['J51'].fill = yellowFill
    ws['D43'].value = 'BONUS WEEK!!!'
    for g in range(8):
        da_char = utils.get_column_letter(g+1)
        ws[f'{da_char}43'].fill = purpleFill


    for j in range(7):
        char = utils.get_column_letter(j+2)
        ws[char + str(11)].value = days_of_week[j]
    for i in range(7):
        if (i == 6):
            ws[f'A{(14 + (i * 5))}'].value = f'Week {i + 1}'
            ws[f'A{(14 + (i * 5))}'].fill = redbrownFill
            ws[f'A{(15 + (i * 5))}'].value = f'Planned'
            ws[f'A{(15 + (i * 5))}'].fill = lightgreenFill
            ws[f'A{(16 + (i * 5))}'].value = f'Actual'
            ws[f'A{(17 + (i * 5))}'].value = f'Exercise'
            ws[f'A{(18 + (i * 5))}'].value = f'Weight'
            ws[f'A{(18 + (i * 5))}'].fill = dirtyyellowFill
            for t in range(7):
                ze_char = utils.get_column_letter(t+2)
                ws[f'{ze_char}{(18 + (i * 5))}'].fill = dirtyyellowFill
            if (start_date == 1):
                if ((month_num - 1) == 2):
                    if isLeap(year) == True:
                        ws['J13'].value = f'Be Healthier on {num_vs_mon_dict[month_num - 1]} 29, {year}'
                    else:
                        ws['J13'].value = f'Be Healthier on {num_vs_mon_dict[month_num - 1]} {day_in_month_dict_num[month_num]}, {year}'
                elif (month_num == 1):
                    ws['J13'].value = f'Be Healthier on Dec 31, {year - 1}'
                else:
                    ws['J13'].value = f'Be Healthier on {num_vs_mon_dict[month_num - 1]} {day_in_month_dict_num[month_num]}, {year}'
            else:
                ws['J13'].value = f'Be Healthier on {num_vs_mon_dict[month_num]} {start_date}, {year}'
        else:
            ws[f'A{(12 + (i * 5))}'].value = f'Week {i + 1}'
            ws[f'A{(12 + (i * 5))}'].fill = redbrownFill
            ws[f'A{(13 + (i * 5))}'].value = f'Planned'
            ws[f'A{(13 + (i * 5))}'].fill = lightgreenFill
            ws[f'A{(14 + (i * 5))}'].value = f'Actual'
            ws[f'A{(15 + (i * 5))}'].value = f'Exercise'
            ws[f'A{(16 + (i * 5))}'].value = f'Weight'
            ws[f'A{(16 + (i * 5))}'].fill = dirtyyellowFill
            for t in range(7):
                ze_char = utils.get_column_letter(t+2)
                ws[f'{ze_char}{(16 + (i * 5))}'].fill = dirtyyellowFill
        for f in range(7):
            if start_date == og_start_date:
                og_start_date = 0
            else:
                start_date += 1
            temp_date = start_date - 1

            temp_month = month_num
            temp_year = year

            if temp_date == 0:
                temp_month -= 1
                if temp_month == 0:
                    temp_month = 12
                    temp_year -= 1
                temp_date = day_in_month_dict_num[temp_month]


            if (i == 6):
                if month_num == 2:
                    if (isLeap(year)) == True:
                        if start_date == 30:
                            month_num += 1
                            start_date = 1
                    else:
                        if start_date == 29:
                            month_num += 1
                            start_date = 1
                if start_date == 31:
                    if day_in_month_dict_num[month_num] == 30:
                        month_num += 1
                        start_date = 1
                if start_date == 32:
                    if month_num == 12:
                        month_num = 1
                        year += 1
                    else:
                        month_num += 1
                    start_date = 1
                char = utils.get_column_letter(f + 2)
                ws[char + str(14 + (i * 5))].value = f'{start_date}-{num_vs_mon_dict[month_num]}'
                ws[char + str(15 + (i * 5))].value = f'{the_list[counter]}'
                if (f'{temp_month}/{temp_date}/{temp_year}' in holidays.US()) or (f'{temp_date}-{num_vs_mon_dict[temp_month]}' in vaca_list):
                    if (f'{month_num}/{start_date}/{year}' in holidays.US()) or (f'{start_date}-{num_vs_mon_dict[month_num]}' in vaca_list):
                        ws[char + str(15 + (i * 5))].fill = lightblueFill
                        ws[char + str(15 + (i * 5))].value = '8-24hr'
                    else:
                        if (datetime.datetime(year, month_num, start_date).weekday() == 5) or (datetime.datetime(year, month_num, start_date).weekday() == 6):
                            ws[char + str(15 + (i * 5))].fill = lightgreenFill
                        else:
                            ws[char + str(15 + (i * 5))].fill = purpleFill
                            ws[char + str(15 + (i * 5))].value = '12-24hr'
                elif (f'{month_num}/{start_date}/{year}' in holidays.US()) or (f'{temp_date}-{num_vs_mon_dict[temp_month]}' in vaca_list):
                    ws[char + str(15 + (i * 5))].fill = lightblueFill
                    ws[char + str(15 + (i * 5))].value = '8-24hr'

                elif counter == 2:
                    temp_date = start_date - 2

                    temp_month = month_num
                    temp_year = year

                    if temp_date == 0:
                        temp_month -= 1
                        if temp_month == 0:
                            temp_month = 12
                            temp_year -= 1
                        temp_date = day_in_month_dict_num[temp_month]
                    if temp_date == -1:
                        temp_month -= 1
                        if temp_month == 0:
                            temp_month = 12
                            temp_year -= 1
                        temp_date = day_in_month_dict_num[temp_month] - 1
                    if (f'{temp_month}/{temp_date}/{temp_year}' in holidays.US()) or (f'{temp_date}-{num_vs_mon_dict[temp_month]}' in vaca_list):
                        ws[char + str(15 + (i * 5))].fill = lightgreenFill
                    else:
                        ws[char + str(15 + (i * 5))].fill = purpleFill
                elif counter > 2:
                    if the_list[counter] == '***':
                        ws[char + str(15 + (i * 5))].fill = yellowFill
                    elif char == 'C':
                        temp_date = start_date + 1
                        temp_month = month_num
                        temp_year = year
                        if temp_month == 2:
                            if (isLeap(temp_year)) == True:
                                if temp_date == 30:
                                    temp_month += 1
                                    temp_date = 1
                            else:
                                if temp_date == 29:
                                    temp_month += 1
                                    temp_date = 1

                        if temp_date == 31:
                            if day_in_month_dict_num[temp_month] == 30:
                                temp_month += 1
                                temp_date = 1
                        if temp_date == 32:
                            temp_month += 1
                            if temp_month == 13:
                                temp_month = 1
                                temp_year += 1
                            temp_date = 1
                        true_check = True
                        for y in range(3):
                            if (f'{temp_month}/{temp_date}/{temp_year}' in holidays.US()) or (
                                    f'{temp_date}-{num_vs_mon_dict[temp_month]}' in vaca_list):
                                true_check = False
                                break
                            temp_date += 1
                            if temp_month == 2:
                                if (isLeap(temp_year)) == True:
                                    if temp_date == 30:
                                        temp_month += 1
                                        temp_date = 1
                                else:
                                    if temp_date == 29:
                                        temp_month += 1
                                        temp_date = 1

                            if temp_date == 31:
                                if day_in_month_dict_num[temp_month] == 30:
                                    temp_month += 1
                                    temp_date = 1
                            if temp_date == 32:
                                temp_month += 1
                                if temp_month == 13:
                                    temp_month = 1
                                    temp_year += 1
                                temp_date = 1
                        if true_check == True:
                            ws[char + str(15 + (i * 5))].fill = purpleFill
                        else:
                            ws[char + str(15 + (i * 5))].fill = lightgreenFill
                    else:
                        ws[char + str(15 + (i * 5))].fill = lightgreenFill
                else:
                    ws[char + str(15 + (i * 5))].fill = lightgreenFill
            else:
                if month_num == 2:
                    if (isLeap(year)) == True:
                        if start_date == 30:
                            month_num += 1
                            start_date = 1
                    else:
                        if start_date == 29:
                            month_num += 1
                            start_date = 1
                if start_date == 31:
                    if day_in_month_dict_num[month_num] == 30:
                        month_num += 1
                        start_date = 1
                if start_date == 32:
                    if month_num == 12:
                        month_num = 1
                        year += 1
                    else:
                        month_num += 1
                    start_date = 1
                char = utils.get_column_letter(f + 2)
                ws[char + str(12 + (i * 5))].value = f'{start_date}-{num_vs_mon_dict[month_num]}'
                ws[char + str(13 + (i * 5))].value = f'{the_list[counter]}'
                if (f'{temp_month}/{temp_date}/{temp_year}' in holidays.US()) or (f'{temp_date}-{num_vs_mon_dict[temp_month]}' in vaca_list):
                    if (f'{month_num}/{start_date}/{year}' in holidays.US()) or (f'{start_date}-{num_vs_mon_dict[month_num]}' in vaca_list):
                        ws[char + str(13 + (i * 5))].fill = lightblueFill
                        ws[char + str(13 + (i * 5))].value = '8-24hr'
                    else:
                        if (datetime.datetime(year, month_num, start_date).weekday() == 5) or (datetime.datetime(year, month_num, start_date).weekday() == 6):
                            ws[char + str(13 + (i * 5))].fill = lightgreenFill
                        else:
                            ws[char + str(13 + (i * 5))].fill = purpleFill
                            ws[char + str(13 + (i * 5))].value = '12-24hr'
                elif (f'{month_num}/{start_date}/{year}' in holidays.US()) or (f'{start_date}-{num_vs_mon_dict[month_num]}' in vaca_list):
                    ws[char + str(13 + (i * 5))].fill = lightblueFill
                    ws[char + str(13 + (i * 5))].value = '8-24hr'
                elif counter == 2:
                    temp_date = start_date - 2

                    temp_month = month_num
                    temp_year = year

                    if temp_date == 0:
                        temp_month -= 1
                        if temp_month == 0:
                            temp_month = 12
                            temp_year -= 1
                        temp_date = day_in_month_dict_num[temp_month]
                    if temp_date == -1:
                        temp_month -= 1
                        if temp_month == 0:
                            temp_month = 12
                            temp_year -= 1
                        temp_date = day_in_month_dict_num[temp_month] - 1
                    if (f'{temp_month}/{temp_date}/{temp_year}' in holidays.US()) or (
                            f'{temp_date}-{num_vs_mon_dict[temp_month]}' in vaca_list):
                        ws[char + str(13 + (i * 5))].fill = lightgreenFill
                    else:
                        ws[char + str(13 + (i * 5))].fill = purpleFill
                elif counter > 2:
                    if the_list[counter] == '***':
                        ws[char + str(13 + (i * 5))].fill = yellowFill
                    elif char == 'C':
                        temp_date = start_date + 1
                        temp_month = month_num
                        temp_year = year
                        if temp_month == 2:
                            if (isLeap(temp_year)) == True:
                                if temp_date == 30:
                                    temp_month += 1
                                    temp_date = 1
                            else:
                                if temp_date == 29:
                                    temp_month += 1
                                    temp_date = 1

                        if temp_date == 31:
                            if day_in_month_dict_num[temp_month] == 30:
                                temp_month += 1
                                temp_date = 1
                        if temp_date == 32:
                            temp_month += 1
                            if temp_month == 13:
                                temp_month = 1
                                temp_year += 1
                            temp_date = 1
                        true_check = True
                        for y in range(3):
                            if (f'{temp_month}/{temp_date}/{temp_year}' in holidays.US()) or (f'{temp_date}-{num_vs_mon_dict[temp_month]}' in vaca_list):
                                true_check = False
                                break
                            temp_date += 1
                            if temp_month == 2:
                                if (isLeap(temp_year)) == True:
                                    if temp_date == 30:
                                        temp_month += 1
                                        temp_date = 1
                                else:
                                    if temp_date == 29:
                                        temp_month += 1
                                        temp_date = 1

                            if temp_date == 31:
                                if day_in_month_dict_num[temp_month] == 30:
                                    temp_month += 1
                                    temp_date = 1
                            if temp_date == 32:
                                temp_month += 1
                                if temp_month == 13:
                                    temp_month = 1
                                    temp_year += 1
                                temp_date = 1
                        if true_check == True:
                            ws[char + str(13 + (i * 5))].fill = purpleFill
                        else:
                            ws[char + str(13 + (i * 5))].fill = lightgreenFill
                    else:
                        ws[char + str(13 + (i * 5))].fill = lightgreenFill
                else:
                    ws[char + str(13 + (i * 5))].fill = lightgreenFill

            counter += 1

    wb.save(f"{First_name}-{Last_name} 6 week plan.xlsx")
if __name__ == "__main__":
    main()
